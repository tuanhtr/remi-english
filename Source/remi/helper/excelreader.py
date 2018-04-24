# -*- coding: utf-8 -*-
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import os.path
import csv


class ExcelReader:
    """
    Use to read block of data from excel
    """
    filename = None
    wb = None

    def __init__(self):
        self.filename = None
        self.wb = None

    def open_csv(self, filepath):
        """
        Open CSV file.
        @param filepath:
        @return:
        """
        if not os.path.isfile(filepath):
            return False
        _, file_extension = os.path.splitext(filepath)

        if file_extension.lower() != '.csv':
            return False

        self.filename = filepath
        self.wb = None

        f = open(filepath, encoding='Shift_JIS')
        try:
            # csv.register_dialect('colons', delimiter=':')
            # reader = csv.reader(f, dialect='colons')
            reader = csv.reader(f)
            self.wb = Workbook()
            ws = self.wb.worksheets[0]
            ws.title = "csv"
            for row_index, row in enumerate(reader):
                for column_index, cell in enumerate(row):
                    column_letter = get_column_letter((column_index + 1))
                    ws.cell('%s%s' % (column_letter, (row_index + 1))).value = cell
        finally:
            f.close()

        return True

    def open(self, filename):
        """
        Open excel file
        @param filename: File name
        @return: True if have no problem
        """
        self.filename = filename
        self.wb = None
        if os.path.isfile(filename):
            try:
                self.wb = load_workbook(filename)
                return True
            except:
                return False
        return False

    def read_csv_block(self, start_col, end_col, start_row=1):
        """
        As read_block, but for CSV data.
        @param start_col:
        @param end_col:
        @param start_row:
        @return:
        """
        return self.read_block(0, start_col, end_col, start_row)

    def read_block(self, sheet_name, start_col, end_col, start_row=1):
        """
        Read a block of data form sheet name. Read until start col have no data.
        @param sheet_name: Sheet name
        @param start_col: Start col
        @param end_col: End col
        @param start_row: Start row
        @return: 2 ways array of data
        """
        if not self.wb:
            return None
        if isinstance(sheet_name, int):
            ws = self.wb.worksheets[sheet_name]
        else:
            ws = self.wb[sheet_name]
        ret = []
        row = start_row
        c = ws.cell(row=row, column=start_col).value
        while c and c != '':
            rec = []
            for i in range(start_col, end_col + 1):
                rec.append(ws.cell(row=row, column=i).value)
            ret.append(rec)
            row = row + 1
            c = ws.cell(row=row, column=start_col).value

        return ret

    def get_sheet_names(self):
        """
        Get list of sheet name
        @rtype: list
        @return: List of sheet name
        """
        if self.wb:
            return self.wb.get_sheet_names()
        return None
